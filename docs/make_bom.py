"""Erzeugt docs/PiLiDAR_BOM_Specs.xlsx mit mehreren Tabellenblättern:
Stückliste, Komponenten-Specs, GPIO-Belegung, Kondensatoren & Kabel,
Treibervergleich A4988/TMC2209.  Aufruf: python docs/make_bom.py
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "PiLiDAR_BOM_Specs.xlsx")

HEAD = PatternFill("solid", fgColor="1F4E79")
HEADF = Font(bold=True, color="FFFFFF", size=11)
TITLEF = Font(bold=True, size=14, color="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet(wb, name, title, headers, rows, widths):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLEF
    ws.append([])
    ws.append(headers)
    hrow = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hrow, c)
        cell.fill = HEAD; cell.font = HEADF; cell.alignment = WRAP; cell.border = BORDER
    for r in rows:
        ws.append(r)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(ws.max_row, c)
            cell.alignment = WRAP; cell.border = BORDER
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(hrow + 1, 1)
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # --- Stückliste ---
    sheet(wb, "Stückliste", "PiLiDAR 2.0 — Stückliste (BOM)",
          ["Pos", "Komponente", "Spezifikation", "Menge", "Richtpreis", "Zweck / Hinweis"],
          [
            [1, "Raspberry Pi 4 (8 GB)", "BCM2711, 4×A72 1,8 GHz", 1, "~75 €", "Server + Rohdatenerfassung"],
            [2, "LDROBOT STL27L LiDAR", "DTOF, 360°, USB-Controllerboard (CP2102)", 1, "~150 €", "Messsensor, UART 921600 über USB"],
            [3, "Schrittmotor NEMA17 42-23", "1,8°, ~1,2 A/Phase, 17 Ncm", 1, "~20 €", "Dreht die Scan-Ebene"],
            [4, "Treiber A4988 (Ist-Zustand)", "1/16 Microstep, VREF-Poti", 1, "~8 €", "Modus A/B; günstig, vorhanden"],
            [5, "Treiber TMC2209 (Upgrade)", "bis 1/256, StealthChop2, UART", 1, "~12 €", "leiser/glatter für Modus B"],
            [6, "Planetengetriebe (3D-Druck)", "Verhältnis 1+38/14 ≈ 3,714:1", 1, "Druck", "Untersetzung, Drehmoment"],
            [7, "STL27L-Halter (3D-Druck)", "PETG, parametrisch (hardware/*.scad)", 1, "Druck", "Seitenmontage, Scan-Ebene vertikal"],
            [8, "Elektrolytkondensator", "100 µF, ≥25 V, low-ESR", 2, "~1 €", "an VMOT/GND jedes Treibers (Pflicht)"],
            [9, "Keramikkondensator", "0,1 µF, X7R, 50 V", 3, "<1 €", "VDD/Logik-Entkopplung"],
            [10, "Elektrolytkondensator", "10 µF", 2, "<1 €", "Stützung 5-V-Schiene Pi/LiDAR (optional)"],
            [11, "Netzteil Pi", "USB-C 5 V / ≥3 A", 1, "~12 €", "separate Pi-Versorgung"],
            [12, "Netzteil Motor", "12 V / ≥3 A", 1, "~12 €", "separate Motorversorgung (VMOT)"],
            [13, "Litze Motor", "18–20 AWG, verdrillt", "~2 m", "~3 €", "Motorphasen A±/B±"],
            [14, "Litze Signal", "24–28 AWG", "~3 m", "~3 €", "STEP/DIR/EN/MS, Masse"],
            [15, "USB-Kabel (geschirmt)", "A→passend, kurz", 1, "~5 €", "LiDAR an Pi-USB"],
            [16, "Ferritkern (Klappferrit)", "für Signal-/USB-Leitung", 2, "~2 €", "EMV-Entstörung"],
            [17, "Schrauben-Set", "M3 (Rotor), M2,5 (LiDAR)", 1, "~5 €", "Montage"],
          ],
          [5, 26, 34, 8, 11, 40])

    # --- Komponenten-Specs (verifiziert) ---
    sheet(wb, "Komponenten-Specs", "Komponenten-Spezifikationen (Datenblätter)",
          ["Gerät", "Parameter", "Wert", "Quelle"],
          [
            ["STL27L", "Messbereich", "0,03–25 m (80 % Reflexion)", "STL-27L Datasheet V0.3"],
            ["STL27L", "Genauigkeit", "±15 mm (0,03–2 m), ±20 mm (2–8 m), ±30 mm (>8 m)", "Datasheet V0.3"],
            ["STL27L", "Messrate", "21.600 Hz (fix)", "Datasheet V0.3"],
            ["STL27L", "Scanfrequenz", "6 / 10 / 13 Hz (PWM-geregelt)", "Datasheet V0.3"],
            ["STL27L", "Winkelauflösung", "0,167° @ 10 Hz", "Datasheet V0.3"],
            ["STL27L", "Maße (L×B×H)", "54,00 × 46,29 × 34,80 mm", "Datasheet V0.3"],
            ["STL27L", "Versorgung", "4,5 / 5,0 / 5,5 V; 290 mA (Start 540 mA)", "Datasheet V0.3"],
            ["STL27L", "Schnittstelle", "UART 921600, 3,3 V; ZH1.5T-4P (Tx,PWM,GND,VCC)", "Datasheet V0.3"],
            ["STL27L", "PWM-Steuersignal", "20–50 kHz, 3,3 V, 40 % → 10 Hz", "Datasheet V0.3"],
            ["NEMA17 42-23", "Schrittwinkel", "1,8° (200 Schritte/U)", "Hersteller/Components101"],
            ["NEMA17 42-23", "Strom/Phase", "~1,2 A (Variante 1,68 A)", "Components101"],
            ["NEMA17 42-23", "Haltemoment", "~17 Ncm", "PiLiDAR / Hersteller"],
            ["A4988", "VMOT / Logik", "8–35 V / 3–5,5 V", "Pololu A4988"],
            ["A4988", "Strom", "~1 A/Phase (2 A mit Kühlung); Pflicht 100 µF an VMOT", "Pololu A4988"],
            ["A4988", "Microstepping", "bis 1/16 (MS1/2/3)", "Pololu A4988"],
            ["TMC2209", "VMOT / Strom", "4,75–28 V / bis 2 A RMS", "Trinamic TMC2209"],
            ["TMC2209", "Microstepping", "bis 1/256, StealthChop2 (leise)", "Trinamic TMC2209"],
            ["TMC2209", "Steuerung", "STEP/DIR + optional UART (PDN)", "Trinamic TMC2209"],
            ["Raspberry Pi 4", "Logikpegel / UART", "3,3 V; UART bis 921600", "RPi Doku"],
            ["Raspberry Pi 4", "Hardware-PWM", "GPIO18=PWM0, GPIO19=PWM1 (dtoverlay pwm-2chan)", "RPi Doku"],
            ["Raspberry Pi 4", "Versorgung", "5 V USB-C, ≥3 A", "RPi Doku"],
          ],
          [16, 22, 52, 26])

    # --- GPIO-Belegung ---
    sheet(wb, "GPIO-Belegung", "GPIO-Belegung (BCM) & Schnittstellen",
          ["Signal", "GPIO/Pin", "Ziel", "Hinweis"],
          [
            ["STEP", "GPIO19 / Pin 35", "Treiber STEP", "= Hardware-PWM-Kanal 1 (Modus B!)"],
            ["DIR", "GPIO26 / Pin 37", "Treiber DIR", "Drehrichtung"],
            ["MS1", "GPIO5 / Pin 29", "Treiber MS1", "Microstep-Wahl"],
            ["MS2", "GPIO6 / Pin 31", "Treiber MS2", "Microstep-Wahl"],
            ["MS3", "GPIO13 / Pin 33", "Treiber MS3", "Microstep-Wahl (1/16)"],
            ["GND", "Pin 6/9/14/…", "Sternmasse", "gemeinsam Pi/Treiber/NT"],
            ["VDD (Logik)", "5 V Pin 2/4", "Treiber VDD", "optional; RST↔SLP brücken"],
            ["UART (nur TMC2209)", "GPIO14/15", "PDN_UART (1 kΩ)", "Strom in Software statt VREF"],
            ["LiDAR", "USB-A", "STL27L (CP2102)", "/dev/ttyUSB0 – kein Pegelwandler nötig"],
          ],
          [20, 18, 22, 42])

    # --- Kondensatoren & Kabel ---
    sheet(wb, "Kondensatoren & Kabel", "Kondensatoren, Kabel & Schirmung",
          ["Element", "Wert/Typ", "Platzierung", "Funktion"],
          [
            ["Elko", "100 µF, ≥25 V, low-ESR", "direkt an VMOT/GND des Treibers", "PFLICHT: LC-Spitzen verhindern"],
            ["Keramik", "0,1 µF, X7R", "an VDD/GND des Treibers (<5 mm)", "HF-Entkopplung Logik"],
            ["Elko", "10 µF", "an 5-V-Schiene Pi/LiDAR", "Stützung (optional)"],
            ["Motorlitze", "18–20 AWG, verdrillt", "Treiber→Motor (A±/B±)", "Strom; verdrillt gegen EMV"],
            ["Signallitze", "24–28 AWG", "Pi→Treiber (STEP/DIR/MS)", "geringe Ströme"],
            ["Schirm", "einseitig auf GND (Treiberseite)", "Signalleitungen", "Brummschleifen vermeiden"],
            ["Ferritkern", "Klappferrit", "USB-/Signalleitung", "HF-Störungen dämpfen"],
            ["Masseführung", "Sternpunkt", "ein gemeinsamer GND-Punkt", "stabile Logikpegel"],
          ],
          [16, 26, 36, 36])

    # --- Treibervergleich ---
    sheet(wb, "Treibervergleich", "A4988 (Ist) vs. TMC2209 (Upgrade)",
          ["Kriterium", "A4988", "TMC2209", "Relevanz für PiLiDAR"],
          [
            ["Microstepping", "bis 1/16", "bis 1/256 (+MicroPlyer)", "feinere, glattere Drehung (Modus B)"],
            ["Geräusch", "hörbar", "nahezu lautlos (StealthChop2)", "ruhigerer Scan"],
            ["Stromeinstellung", "VREF-Poti", "VREF oder UART (Software)", "Komfort/Wiederholbarkeit"],
            ["VMOT", "8–35 V", "4,75–28 V", "12-V-Betrieb bei beiden ok"],
            ["Schrittverlust-Risiko", "höher", "geringer (sanfter)", "weniger Loop-Closure-Fehler"],
            ["Verkabelung", "STEP/DIR/EN/MS", "STEP/DIR/EN (+UART opt.)", "weitgehend identisch"],
            ["100 µF an VMOT", "Pflicht", "Pflicht", "beide"],
            ["Preis", "~8 €", "~12 €", "geringer Aufpreis"],
          ],
          [22, 24, 30, 36])

    wb.save(OUT)
    print("ok", OUT)


if __name__ == "__main__":
    build()
